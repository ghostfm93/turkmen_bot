import os
import openpyxl
from openpyxl import load_workbook
context = {
     'name_latin': 'Oleg Mongol',
     'name_cyrill': 'Олег Монгол',
     'birth_date': '21.02.2021',
     'passport_number': 'A1234567',
     'passport_given': '01.01.2023',
     'passport_ends': '01.01.2025',
     'registration_address': 'г. Витебск, ул. Хуевая',
     'telephone': '+375333069771',
}

def make_dcm_petition():
     '''делаем ходатайство в ОГИМ'''
     wb = load_workbook('test.xlsx')
     sheet = wb['стр.1']
     indexes = {
          'name_latin' : {14 : range(2,113,4)},
          'name_cyrill' : {16 : range(2,113,4)},
          'birth_date' : {18 : [17,24,41]},
          'passport_number' : {26 : [10, 59, 64, 69, 74, 79, 84, 89]},
          'passport_given' : {27 : [12,22,46]},
          'passport_ends' : {27 : [76,86,110]},
          'registration_address' : {35 : [13,]},
          'telephone' : {37 : [2,]},
          'ogim' : {5 : [43,]},
          'number' : {5 : [6,]},
     }


     # for cat, ind in indexes.items():
     #      curr_val = context[cat]
     #      if cat in ['birth_date', 'passport_given', 'passport_ends']:
     #           curr_val = curr_val.split('.')
     #      for string, rows in ind.items():
     #           for num, i in enumerate(rows):
     #                if len(rows) == 1:
     #                     sheet.cell(row = string, column = i).value = curr_val
     #                elif num < len(curr_val):
     #                     if isinstance(curr_val[num], str):
     #                          sheet.cell(row = string, column = i).value = curr_val[num].upper()
     #                     else:
     #                          sheet.cell(row = string, column = i).value = curr_val[num]
     #
     # print(wb.save('test1.xlsx'))

     for i in range(1,200):
          for j in range(1,200):
               if sheet.cell(row = i, column = j).value != None:
                    print(i, j, sheet.cell(row = i, column = j).value)

make_dcm_petition()
